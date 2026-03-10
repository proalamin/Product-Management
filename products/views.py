from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import load_workbook

from .forms import ApprovedFilterForm, ExcelUploadForm
from .models import Product


def process_excel(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = ['product_id', 'name', 'category', 'price', 'quantity']
    header_lower = [str(h).strip().lower() if h else '' for h in headers]

    if header_lower[:5] != expected:
        wb.close()
        raise ValueError('Invalid Excel format. Expected columns: product_id, name, category, price, quantity')

    inserted = 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        obj, created = Product.objects.update_or_create(
            product_id=int(row[0]),
            defaults={
                'name': str(row[1]).strip(),
                'category': str(row[2]).strip(),
                'price': float(row[3]),
                'quantity': int(row[4]),
                'status': 'Draft',
            }
        )
        if created:
            inserted += 1
        else:
            updated += 1

    wb.close()
    return inserted, updated


def draft_list(request):
    form = ExcelUploadForm()
    products = Product.objects.all()
    paginator = Paginator(products, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'products/draft.html', {
        'form': form,
        'page_obj': page_obj,
    })


def upload_excel(request):
    if request.method != 'POST':
        return redirect('draft_list')

    form = ExcelUploadForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            inserted, updated = process_excel(form.cleaned_data['file'])
            messages.success(request, f'File uploaded successfully! {inserted} inserted, {updated} updated.')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error processing file: {e}')
    else:
        messages.error(request, 'Please select a valid .xlsx file.')

    return redirect('draft_list')


def approve_product(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)
    product.status = 'Approved'
    product.save()
    messages.success(request, f'Product "{product.name}" (ID: {product.product_id}) approved successfully!')
    return redirect('draft_list')


def approved_products(request):
    form = ApprovedFilterForm(request.GET or None)
    queryset = Product.objects.filter(status='Approved')
    per_page = '10'

    if form.is_valid():
        search = form.cleaned_data.get('search')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        per_page = form.cleaned_data.get('per_page') or '10'

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(category__icontains=search)
            )
        if date_from:
            queryset = queryset.filter(last_updated__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(last_updated__date__lte=date_to)

    paginator = Paginator(queryset, int(per_page))
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'products/approved.html', {
        'form': form,
        'page_obj': page_obj,
    })
